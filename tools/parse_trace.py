#!/usr/bin/env python3
"""Small, step-by-step ATOM trace parser.

Step 1: find the decode warmup window in the capture trace that corresponds to
the first decode event in the run trace.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import os
import re
from collections.abc import Iterable
from glob import glob
from typing import Any

SPECIAL_KERNEL_LAUNCH_NAMES = {"hipmemcpyasync"}


def load_events(path: str) -> list[dict[str, Any]]:
    opener = gzip.open if path.endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", errors="replace") as f:
        return json.load(f).get("traceEvents", [])


def event_end(event: dict[str, Any]) -> float:
    return float(event.get("ts", 0.0)) + float(event.get("dur", 0.0))


def is_kernel_launch(name: str) -> bool:
    normalized = name.lower()
    return (
        "launch" in normalized and "kernel" in normalized
    ) or normalized in SPECIAL_KERNEL_LAUNCH_NAMES


def short(text: Any, limit: int = 80) -> str:
    value = str(text)
    return value if len(value) <= limit else value[: limit - 3] + "..."


def model_name_from_trace(path: str) -> str | None:
    base = os.path.basename(path)
    if "_ts_" not in base:
        return None
    prefix = base.split("_ts_", 1)[0]
    if prefix.startswith("capture_graph_"):
        prefix = prefix[len("capture_graph_") :]
    return prefix or None


def find_legacy_capture_trace(run_trace: str) -> str | None:
    """Locate the single whole-phase capture trace, if one is lying around.

    Superseded by the per-(bs, q-bucket) files under ``capture_traces/``; kept
    so traces captured before the split still parse.
    """
    model_name = model_name_from_trace(run_trace)
    if not model_name:
        return None
    trace_dir = os.path.dirname(run_trace) or "."
    pattern = os.path.join(trace_dir, f"capture_graph_{model_name}_ts_*.pt.trace.json*")
    candidates = sorted(glob(pattern), key=os.path.getmtime, reverse=True)
    run_abs = os.path.abspath(run_trace)
    for candidate in candidates:
        if os.path.abspath(candidate) != run_abs:
            return candidate
    return None


def resolve_capture_trace(run_trace: str, graph_bs: int, q_len: int | None) -> str:
    """Return the capture trace holding the graph this decode replayed.

    Capture is written one file per (batch size, q-bucket) into
    ``{run_trace_dir}/capture_traces/``, so the batch size now selects the
    *file* instead of a span inside one combined trace.
    """
    capture_dir = os.path.join(os.path.dirname(run_trace) or ".", "capture_traces")
    if not os.path.isdir(capture_dir):
        legacy = find_legacy_capture_trace(run_trace)
        if legacy is not None:
            return legacy
        raise RuntimeError(
            f"No {capture_dir} directory and no legacy capture trace next to "
            f"{run_trace}. Re-run with --mark-trace, or pass --capture-trace."
        )

    def matching(q: str) -> list[str]:
        pattern = os.path.join(capture_dir, f"bs_{graph_bs}_q_{q}_rank*.json.gz")
        return sorted(glob(pattern))

    # Prefer the exact q-bucket; fall back to any bucket for this batch size so
    # a label without a usable tok= field still resolves when it is unambiguous.
    candidates = matching(str(q_len)) if q_len is not None else []
    if not candidates:
        candidates = matching("*")
    if len(candidates) == 1:
        return candidates[0]
    if candidates:
        names = ", ".join(os.path.basename(path) for path in candidates)
        raise RuntimeError(
            f"Multiple capture traces match bs={graph_bs} in {capture_dir} "
            f"({names}); pass --capture-trace to choose one."
        )
    available = sorted(
        os.path.basename(path)
        for path in glob(os.path.join(capture_dir, "bs_*_rank*.json.gz"))
    )
    raise RuntimeError(
        f"No capture trace for bs={graph_bs} in {capture_dir}. "
        f"Present: {', '.join(available) if available else '(none)'}"
    )


def find_first_decode(events: list[dict[str, Any]]) -> dict[str, Any]:
    decodes = sorted(
        [
            event
            for event in events
            if event.get("ph") == "X"
            and event.get("cat") == "gpu_user_annotation"
            and str(event.get("name", "")).startswith("decode[")
        ],
        key=lambda event: event["ts"],
    )
    if not decodes:
        raise RuntimeError("No decode gpu_user_annotation found in run trace.")
    return decodes[0]


def decode_batch_sizes(decode_event: dict[str, Any]) -> tuple[int, int]:
    """Return ``(scheduled_bs, graph_bs)`` from a decode label.

    ``build_run_label`` writes ``bs=<real>/<graph>`` when the CUDAGraph replays
    a padded batch. Capture files are named after the *graph* size, so the two
    must not be conflated when picking one.
    """
    match = re.search(r"bs=(\d+)(?:/(\d+))?", str(decode_event.get("name", "")))
    if not match:
        raise RuntimeError(
            f"Could not parse batch size from {decode_event.get('name')!r}"
        )
    scheduled = int(match.group(1))
    return scheduled, int(match.group(2)) if match.group(2) else scheduled


def decode_query_len(decode_event: dict[str, Any], scheduled_bs: int) -> int | None:
    """Query tokens per request — 1, or ``mtp_k + 1`` under spec decode.

    Selects the q-bucket among the capture files. ``None`` when the label
    carries no ``tok=`` field that divides evenly by the batch size.
    """
    match = re.search(r"tok=(\d+)", str(decode_event.get("name", "")))
    if not match or scheduled_bs <= 0:
        return None
    q_len, remainder = divmod(int(match.group(1)), scheduled_bs)
    return q_len if remainder == 0 and q_len > 0 else None


def find_cpu_capture_graphs(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        [
            event
            for event in events
            if event.get("ph") == "X"
            and event.get("cat") == "user_annotation"
            and str(event.get("name", "")).startswith("capture_graph_bs_")
        ],
        key=lambda event: event["ts"],
    )


def find_capture_graph_for_bs(
    capture_events: list[dict[str, Any]], batch_size: int
) -> dict[str, Any]:
    graphs = find_cpu_capture_graphs(capture_events)
    if not graphs:
        raise RuntimeError(
            "No CPU capture_graph_bs_* annotations found in capture trace."
        )
    target_name = f"capture_graph_bs_{batch_size}"
    for graph in graphs:
        if graph.get("name") == target_name:
            return graph
    present = ", ".join(sorted({str(graph.get("name")) for graph in graphs}))
    raise RuntimeError(f"No {target_name} in capture trace; it holds: {present}")


def warmup_window_for_graph(
    capture_events: list[dict[str, Any]], target_graph: dict[str, Any]
) -> tuple[float, float]:
    """Return [previous_capture_graph_end, target_capture_graph_start).

    A per-batch-size capture file holds a single ``capture_graph_bs_*`` span, so
    the window opens at the first event in the file and runs to that span — the
    whole file up to the capture, which is exactly the warmup forward. The scan
    over preceding spans only matters for a legacy combined trace, where every
    batch size shares one file.

    The floor is the earliest event rather than 0.0 so the window duration stays
    a duration; timestamps here are absolute, so 0.0 would report the epoch.
    """
    start = min(
        (float(event["ts"]) for event in capture_events if event.get("ph") == "X"),
        default=0.0,
    )
    for graph in find_cpu_capture_graphs(capture_events):
        if graph is target_graph:
            return start, float(target_graph["ts"])
        start = max(start, event_end(graph))
    raise RuntimeError("Target capture graph was not in capture graph list.")


def count_events_in_window(
    events: list[dict[str, Any]], start: float, end: float
) -> dict[str, int]:
    counts = {"duration": 0, "user_annotation": 0, "cuda_runtime": 0, "kernel": 0}
    for event in events:
        if event.get("ph") != "X":
            continue
        ts = float(event.get("ts", 0.0))
        if not (start <= ts < end):
            continue
        counts["duration"] += 1
        cat = event.get("cat")
        if cat in counts:
            counts[cat] += 1
    return counts


def build_correlation_index(
    events: list[dict[str, Any]], start: float, end: float
) -> tuple[dict[Any, dict[str, Any]], dict[Any, dict[str, Any]]]:
    launches: dict[Any, dict[str, Any]] = {}
    kernels: dict[Any, dict[str, Any]] = {}
    for event in events:
        if event.get("ph") != "X":
            continue
        ts = float(event.get("ts", 0.0))
        if not (start <= ts < end):
            continue
        corr = (event.get("args") or {}).get("correlation")
        if corr is None:
            continue
        if event.get("cat") == "cuda_runtime" and is_kernel_launch(
            str(event.get("name", ""))
        ):
            launches.setdefault(corr, event)
        elif event.get("cat") == "kernel":
            kernels.setdefault(corr, event)
    return launches, kernels


def is_profiler_step_tag(name: str) -> bool:
    """kineto's own per-step marker, e.g. ``ProfilerStep#3``.

    It spans the entire profiled window and names nothing in the model, so it
    must never win as a kernel's module tag — it would otherwise be picked for
    every kernel no enclosing ``record_function`` covers (as the only container
    on the GPU side, and as the longest one on the CPU fallback side).
    """
    return name.startswith("ProfilerStep#")


def containing_annotations(
    event: dict[str, Any], annotations: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    start = float(event["ts"])
    end = event_end(event)
    return [
        ann
        for ann in annotations
        if ann.get("pid") == event.get("pid")
        and ann.get("tid") == event.get("tid")
        and float(ann.get("ts", 0.0)) <= start
        and end <= event_end(ann)
        and not is_profiler_step_tag(str(ann.get("name", "")))
    ]


def is_compiled_graph_tag(name: str) -> bool:
    return name.startswith("## Call CompiledFxGraph")


def cpu_fallback_tag_for_compiled_kernel(
    kernel: dict[str, Any],
    launch_by_corr: dict[Any, dict[str, Any]],
    cpu_events: list[dict[str, Any]],
) -> str | None:
    corr = (kernel.get("args") or {}).get("correlation")
    launch = launch_by_corr.get(corr)
    if launch is None:
        return None

    containers = containing_annotations(launch, cpu_events)
    # Include cpu_op parents as well as user annotations, then pick the largest
    # non-CompiledFxGraph container. This maps tiny compiled graph kernels such
    # as FillFunctor copies back to semantic CPU ops like aiter::all_reduce_.
    start = float(launch["ts"])
    end = event_end(launch)
    containers.extend(
        event
        for event in cpu_events
        if event.get("cat") == "cpu_op"
        and event.get("pid") == launch.get("pid")
        and event.get("tid") == launch.get("tid")
        and float(event.get("ts", 0.0)) <= start
        and end <= event_end(event)
    )
    candidates = [
        event
        for event in containers
        if not is_compiled_graph_tag(str(event.get("name", "")))
    ]
    if not candidates:
        return None
    return str(
        max(candidates, key=lambda event: float(event.get("dur", 0.0))).get("name")
    )


def gpu_tag_for_kernel(
    kernel: dict[str, Any],
    gpu_annotations: list[dict[str, Any]],
    launch_by_corr: dict[Any, dict[str, Any]],
    cpu_events: list[dict[str, Any]],
) -> str:
    containers = containing_annotations(kernel, gpu_annotations)
    if not containers:
        return "<no gpu tag>"
    tag = str(
        min(containers, key=lambda event: float(event.get("dur", 0.0))).get("name")
    )
    if is_compiled_graph_tag(tag):
        fallback = cpu_fallback_tag_for_compiled_kernel(
            kernel, launch_by_corr, cpu_events
        )
        if fallback:
            return fallback
    return tag


def build_warmup_mapping(
    capture_events: list[dict[str, Any]], start: float, end: float
) -> list[dict[str, Any]]:
    """Build the internal decode warmup mapping.

    Each row is intentionally minimal:
      - module: resolved CPU/GPU tag name
      - kernel: GPU kernel name
      - stream: GPU stream id

    This mapping is the attribution source for later replay-duration matching;
    it is not meant to be emitted as the final user-facing breakdown.
    """
    launch_by_corr, _ = build_correlation_index(capture_events, start, end)
    cpu_events = [
        event
        for event in capture_events
        if event.get("ph") == "X"
        and start <= float(event.get("ts", 0.0)) < end
        and event.get("cat") in {"user_annotation", "cpu_op"}
    ]
    gpu_annotations = [
        event
        for event in capture_events
        if event.get("ph") == "X"
        and start <= float(event.get("ts", 0.0)) < end
        and event.get("cat") == "gpu_user_annotation"
    ]
    kernels = sorted(
        [
            event
            for event in capture_events
            if event.get("ph") == "X"
            and start <= float(event.get("ts", 0.0)) < end
            and event.get("cat") == "kernel"
        ],
        key=lambda event: event["ts"],
    )
    mapping: list[dict[str, Any]] = []
    for kernel in kernels:
        args = kernel.get("args") or {}
        mapping.append(
            {
                "module": gpu_tag_for_kernel(
                    kernel, gpu_annotations, launch_by_corr, cpu_events
                ),
                "kernel": str(kernel.get("name", "")),
                "stream": args.get("stream"),
            }
        )
    return mapping


def print_first_warmup_mappings(mapping: list[dict[str, Any]], limit: int) -> None:
    print("")
    print(f"First {limit} warmup mappings:")
    print("| # | module/tag | stream | kernel |")
    print("|---:|---|---:|---|")
    for idx, item in enumerate(mapping[:limit]):
        print(
            f"| {idx} | `{short(item['module'], 55)}` | {item['stream']} | "
            f"`{short(item['kernel'], 85)}` |"
        )


def decode_gpu_window(
    run_events: list[dict[str, Any]], decode_event: dict[str, Any]
) -> tuple[float, float]:
    """Return the GPU annotation time range for the selected decode event.

    We intentionally use the GPU-side annotation range here: the final CSV is
    for observed replay GPU kernels. Kernels that fall just outside this range
    are not included in this first formal path.
    """
    external_id = (decode_event.get("args") or {}).get("External id")
    gpu_decodes = [
        event
        for event in run_events
        if event.get("ph") == "X"
        and event.get("cat") == "gpu_user_annotation"
        and (event.get("args") or {}).get("External id") == external_id
    ]
    if not gpu_decodes:
        # Fallback for traces without GPU annotation projection.
        return float(decode_event["ts"]), event_end(decode_event)
    return min(float(event["ts"]) for event in gpu_decodes), max(
        event_end(event) for event in gpu_decodes
    )


def replay_kernels_in_window(
    run_events: list[dict[str, Any]], start: float, end: float
) -> list[dict[str, Any]]:
    return sorted(
        [
            event
            for event in run_events
            if event.get("ph") == "X"
            and event.get("cat") == "kernel"
            and start <= float(event.get("ts", 0.0)) < end
        ],
        key=lambda event: event["ts"],
    )


def remap_streams(replay_kernels: list[dict[str, Any]]) -> dict[Any, int]:
    """Map real replay stream ids to compact 1..N ids by numeric order."""
    streams = sorted(
        {(event.get("args") or {}).get("stream") for event in replay_kernels},
        key=lambda value: (value is None, value),
    )
    return {stream: idx + 1 for idx, stream in enumerate(streams)}


LAYER_RE = re.compile(r"(^|\.)layers\.(\d+)\.")


def module_layer(module: str) -> int | None:
    match = LAYER_RE.search(module)
    return int(match.group(2)) if match else None


def normalize_layer_module(module: str) -> str:
    return re.sub(r"(^|\.)layers\.\d+\.", r"\1layers.*.", module)


def layer_group_label(layers: list[int]) -> str:
    layers = sorted(layers)
    if not layers:
        return "layers <empty>"
    if len(layers) == 1:
        return f"layer {layers[0]}"
    if layers == list(range(layers[0], layers[-1] + 1)):
        return f"layers {layers[0]}-{layers[-1]}"
    if len(layers) <= 8:
        return "layers " + ",".join(str(layer) for layer in layers)
    return f"layers {layers[0]},{layers[1]},...,{layers[-1]} ({len(layers)} layers)"


def warmup_item_owner_layers(warmup_mapping: list[dict[str, Any]]) -> list[int | None]:
    """Assign warmup rows to layer parse windows.

    A few runtime helper kernels, such as maybe_dual_stream_forward, appear as
    non-layer tags between two chunks of the same layer.  They should still be
    consumed while parsing that layer; their original module tag is preserved in
    the final row.
    """
    direct_layers = [module_layer(item["module"]) for item in warmup_mapping]
    owners: list[int | None] = []
    for idx, layer in enumerate(direct_layers):
        if layer is not None:
            owners.append(layer)
            continue

        prev_layer = None
        for prev_idx in range(idx - 1, -1, -1):
            if direct_layers[prev_idx] is not None:
                prev_layer = direct_layers[prev_idx]
                break

        next_layer = None
        for next_idx in range(idx + 1, len(direct_layers)):
            if direct_layers[next_idx] is not None:
                next_layer = direct_layers[next_idx]
                break

        owners.append(
            prev_layer if prev_layer is not None and prev_layer == next_layer else None
        )
    return owners


def warmup_parse_blocks(
    warmup_mapping: list[dict[str, Any]],
) -> list[tuple[int | None, list[tuple[int, dict[str, Any]]]]]:
    owners = warmup_item_owner_layers(warmup_mapping)
    blocks: list[tuple[int | None, list[tuple[int, dict[str, Any]]]]] = []
    for idx, (owner, item) in enumerate(zip(owners, warmup_mapping, strict=True)):
        if not blocks or blocks[-1][0] != owner:
            blocks.append((owner, []))
        blocks[-1][1].append((idx, item))
    return blocks


def primary_replay_stream(replay_kernels: list[dict[str, Any]]) -> Any:
    counts: dict[Any, int] = {}
    first_ts: dict[Any, float] = {}
    for event in replay_kernels:
        stream = (event.get("args") or {}).get("stream")
        counts[stream] = counts.get(stream, 0) + 1
        first_ts.setdefault(stream, float(event.get("ts", 0.0)))
    return max(counts, key=lambda stream: (counts[stream], -first_ts[stream]))


def consume_replay_stream_for_block(
    stream_events: list[dict[str, Any]],
    cursor: int,
    template: list[tuple[int, dict[str, Any]]],
    used_template_positions: set[int],
) -> tuple[int, list[tuple[int, dict[str, Any], dict[str, Any]]]]:
    """Consume one replay stream against one warmup layer block.

    The stream is matched event-by-event to the next compatible warmup kernel in
    this block, skipping template rows that ran on other replay streams.
    """
    rows: list[tuple[int, dict[str, Any], dict[str, Any]]] = []
    template_pos = 0
    pos = cursor
    while pos < len(stream_events):
        replay = stream_events[pos]
        replay_kernel = str(replay.get("name", ""))
        matched_pos = None
        for idx in range(template_pos, len(template)):
            if idx in used_template_positions:
                continue
            if template[idx][1]["kernel"] == replay_kernel:
                matched_pos = idx
                break
        if matched_pos is None:
            break

        used_template_positions.add(matched_pos)
        template_pos = matched_pos + 1
        rows.append((template[matched_pos][0], template[matched_pos][1], replay))
        pos += 1
    return pos, rows


# Ops that launch a different kernel in the eager warmup than in the captured
# graph, keyed by a substring of the replay kernel and mapping to a substring of
# the warmup kernel that stands in for it at the same call site.
#
# aiter's custom all-reduce is the only one today. Its `custom_all_reduce`
# branches on `torch.cuda.is_current_stream_capturing()`: inside capture it
# issues the real `cross_device_reduce_*`, and in the warmup it returns
# `torch.zeros_like(input)` to "mimic the allocation pattern" without
# communicating. So the template holds a FillFunctor where the replay holds the
# reduce — same call site, same position, different kernel.
#
# Only add entries here for a divergence that is deliberate and documented
# upstream; anything else should stay visibly unmatched.
CAPTURE_MODE_KERNEL_SUBSTITUTIONS: tuple[tuple[str, str], ...] = (
    ("cross_device_reduce", "FillFunctor"),
)


def substituted_warmup_kernel(replay_kernel: str) -> str | None:
    """Warmup-kernel substring standing in for *replay_kernel*, if known."""
    for replay_marker, warmup_marker in CAPTURE_MODE_KERNEL_SUBSTITUTIONS:
        if replay_marker in replay_kernel:
            return warmup_marker
    return None


def match_replay_to_warmup(
    replay_kernels: list[dict[str, Any]], warmup_mapping: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Match replay to warmup layer-by-layer without stream-id assumptions.

    Warmup mapping is treated as the semantic operator template.  For each layer
    block, the primary replay stream is consumed first, then the remaining
    streams consume the still-unmatched kernels in the same block.  This avoids
    assuming capture streams and CUDAGraph replay streams are one-to-one.
    """
    if not replay_kernels:
        return []

    replay_by_stream: dict[Any, list[dict[str, Any]]] = {}
    for event in replay_kernels:
        stream = (event.get("args") or {}).get("stream")
        replay_by_stream.setdefault(stream, []).append(event)

    main_stream = primary_replay_stream(replay_kernels)
    other_streams = [
        stream
        for stream in sorted(replay_by_stream, key=lambda value: (value is None, value))
        if stream != main_stream
    ]
    stream_order = [main_stream] + other_streams
    stream_cursors = {stream: 0 for stream in replay_by_stream}
    used_warmup_indices: set[int] = set()
    warmup_owner_layers = warmup_item_owner_layers(warmup_mapping)
    rows: list[dict[str, Any]] = []

    for _, template in warmup_parse_blocks(warmup_mapping):
        used_template_positions: set[int] = set()
        block_matches: list[tuple[int, dict[str, Any], dict[str, Any]]] = []
        for stream in stream_order:
            next_cursor, stream_matches = consume_replay_stream_for_block(
                replay_by_stream[stream],
                stream_cursors[stream],
                template,
                used_template_positions,
            )
            stream_cursors[stream] = next_cursor
            block_matches.extend(stream_matches)

        for warmup_idx, matched, replay in sorted(
            block_matches, key=lambda item: item[0]
        ):
            used_warmup_indices.add(warmup_idx)
            replay_stream = (replay.get("args") or {}).get("stream")
            rows.append(
                {
                    "warmup_index": warmup_idx,
                    "cpu_module": matched["module"],
                    "owner_layer": warmup_owner_layers[warmup_idx],
                    "kernel_name": str(replay.get("name", "")),
                    "stream": replay_stream,
                    "duration_us": float(replay.get("dur", 0.0)),
                    "ts": float(replay.get("ts", 0.0)),
                }
            )

    leftovers: list[dict[str, Any]] = []
    for stream, events in replay_by_stream.items():
        leftovers.extend(events[stream_cursors[stream] :])

    unused_warmup: list[tuple[int, dict[str, Any]]] = [
        (idx, item)
        for idx, item in enumerate(warmup_mapping)
        if idx not in used_warmup_indices
    ]
    unused_warmup_by_kernel: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for idx, item in unused_warmup:
        unused_warmup_by_kernel.setdefault(item["kernel"], []).append((idx, item))
    consumed: set[int] = set()

    def take_exact(replay_kernel: str) -> tuple[int, dict[str, Any] | None]:
        for idx, item in unused_warmup_by_kernel.get(replay_kernel, []):
            if idx not in consumed:
                return idx, item
        return -1, None

    def take_substitute(replay_kernel: str) -> tuple[int, dict[str, Any] | None]:
        marker = substituted_warmup_kernel(replay_kernel)
        if marker is None:
            return -1, None
        for idx, item in unused_warmup:
            if idx not in consumed and marker in item["kernel"]:
                return idx, item
        return -1, None

    for replay in sorted(leftovers, key=lambda event: float(event.get("ts", 0.0))):
        replay_stream = (replay.get("args") or {}).get("stream")
        replay_kernel = str(replay.get("name", ""))
        matched_idx, matched = take_exact(replay_kernel)
        if matched is None:
            matched_idx, matched = take_substitute(replay_kernel)
        if matched_idx >= 0:
            consumed.add(matched_idx)
        rows.append(
            {
                "warmup_index": matched_idx,
                "cpu_module": matched["module"] if matched else "<unmatched>",
                "owner_layer": (
                    warmup_owner_layers[matched_idx] if matched_idx >= 0 else None
                ),
                "kernel_name": replay_kernel,
                "stream": replay_stream,
                "duration_us": float(replay.get("dur", 0.0)),
                "ts": float(replay.get("ts", 0.0)),
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            row["warmup_index"] < 0,
            row["warmup_index"],
        ),
    )


def union_duration(intervals: Iterable[tuple[float, float]]) -> float:
    """Wall-clock covered by *intervals*, counting overlap once.

    Kernels on concurrent streams overlap, so summing their durations measures
    GPU work rather than elapsed time. The union is what a layer actually costs.
    """
    merged: list[list[float]] = []
    for start, end in sorted(intervals):
        if merged and start <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], end)
        else:
            merged.append([start, end])
    return sum(end - start for start, end in merged)


def build_grouped_breakdown_rows(
    rows: list[dict[str, Any]], stream_map: dict[Any, int], print_head: bool = False
) -> list[dict[str, Any]]:
    """Aggregate matched replay rows into layer-structure groups.

    Layer rows are grouped by identical per-layer operator sequence.  Each output
    row for a layer group is the average time for that operator position across
    layers in the group.  Non-layer and unmatched rows are aggregated by
    module/kernel/stream.

    ``time_us`` is therefore per-layer while the decode window it is reported
    against covers every layer, so each row also carries the ``layer_count`` its
    average was taken over — multiply the two to get the share of the forward.
    """
    layer_rows: dict[int, list[dict[str, Any]]] = {}
    non_layer_accum: dict[tuple[str, str, int], tuple[float, int]] = {}
    unmatched_accum: dict[tuple[str, int], tuple[float, int]] = {}

    for row in rows:
        stream_no = stream_map.get(row["stream"], 0)
        if row["cpu_module"] == "<unmatched>":
            key = (row["kernel_name"], stream_no)
            total_us, count = unmatched_accum.get(key, (0.0, 0))
            unmatched_accum[key] = (total_us + row["duration_us"], count + 1)
            continue

        layer = module_layer(row["cpu_module"])
        if layer is None:
            layer = row.get("owner_layer")
        normalized = normalize_layer_module(row["cpu_module"])
        normalized_row = {
            **row,
            "module_pattern": normalized,
            "stream_no": stream_no,
        }
        if layer is None:
            key = (normalized, row["kernel_name"], stream_no)
            total, order_key = non_layer_accum.get(key, (0.0, row["warmup_index"]))
            non_layer_accum[key] = (
                total + row["duration_us"],
                min(order_key, row["warmup_index"]),
            )
        else:
            layer_rows.setdefault(layer, []).append(normalized_row)

    grouped: list[dict[str, Any]] = []

    # Non-layer prologue/epilogue rows are hidden by default. They still
    # contribute to the full-decode denominator.
    if print_head:
        for (module, kernel, stream_no), (
            total_us,
            order_key,
        ) in non_layer_accum.items():
            grouped.append(
                {
                    "layer_group": "non_layer",
                    "layer_count": 1,
                    "module": module,
                    "kernel": kernel,
                    "stream_no": stream_no,
                    "time_us": total_us,
                    "order_key": order_key,
                    "group_order_key": order_key,
                }
            )

    # Layer groups by exact normalized operator sequence.  Stream id is kept as
    # output metadata, but does not decide whether two layers have the same
    # operator layout.
    signature_to_layers: dict[tuple[tuple[str, str], ...], list[int]] = {}
    for layer in sorted(layer_rows):
        items = layer_rows[layer]
        signature = tuple(
            (item["module_pattern"], item["kernel_name"]) for item in items
        )
        signature_to_layers.setdefault(signature, []).append(layer)

    for signature, layers in signature_to_layers.items():
        layer_count = len(layers)
        label = layer_group_label(layers)
        group_order_key = min(layer_rows[layer][0]["warmup_index"] for layer in layers)
        # What one layer of this group actually costs: the union of its kernels'
        # intervals, not the sum of their durations. Under dual-stream execution
        # the side streams (shared_experts, attn.compressor, indexer) run
        # concurrently with the primary one, so summing counts the same elapsed
        # time twice — 15% of the window on a TP=4 V4-Pro decode.
        #
        # Also the denominator of each row's percent_of_current_layer. Per-row
        # time_us stays unmerged, so those shares sum past 100% exactly to the
        # extent the layer's streams overlapped — that excess is the signal.
        layer_total_us = (
            union_duration(
                (item["ts"], item["ts"] + item["duration_us"])
                for layer in layers
                for item in layer_rows[layer]
            )
            / layer_count
        )
        for idx, (module, kernel) in enumerate(signature):
            total = sum(layer_rows[layer][idx]["duration_us"] for layer in layers)
            stream_ids = [layer_rows[layer][idx]["stream_no"] for layer in layers]
            order_key = min(layer_rows[layer][idx]["warmup_index"] for layer in layers)
            stream_no: int | str
            if all(stream_id == stream_ids[0] for stream_id in stream_ids):
                stream_no = stream_ids[0]
            else:
                stream_no = ",".join(
                    str(stream_id) for stream_id in sorted(set(stream_ids))
                )
            grouped.append(
                {
                    "layer_group": label,
                    "layer_count": layer_count,
                    "module": module,
                    "kernel": kernel,
                    "stream_no": stream_no,
                    "time_us": total / layer_count,
                    "layer_total_us": layer_total_us,
                    "order_key": order_key,
                    "group_order_key": group_order_key,
                }
            )
        grouped.append(
            {
                "layer_group": label,
                "layer_count": layer_count,
                "module": "__layer_total__",
                "kernel": "LAYER TOTAL",
                "stream_no": "",
                "time_us": layer_total_us,
                "layer_total_us": layer_total_us,
                "order_key": float("inf"),
                "group_order_key": group_order_key,
            }
        )

    # Kept in the output as a list of what the template could not account for,
    # so a growing tail is visible. Their numbers are suppressed: a kernel with
    # no template counterpart has no layer to be a share of, and adding its time
    # to a breakdown that is already attributed elsewhere only misleads.
    #
    # The label carries the kernel count so the tail's size is visible in the
    # sheet itself — the rows below collapse repeats, so counting them does not
    # give it. One constant label also keeps the group cell merging as one run.
    unmatched_total = sum(count for _, count in unmatched_accum.values())
    unmatched_label = f"unmatched: {unmatched_total} kernels (should be ignored)"
    for (kernel, stream_no), (total_us, _count) in unmatched_accum.items():
        grouped.append(
            {
                "layer_group": unmatched_label,
                "layer_count": 1,
                "module": "<unmatched>",
                "kernel": kernel,
                "stream_no": stream_no,
                "time_us": total_us,
                "suppress_numbers": True,
                "order_key": float("inf"),
                "group_order_key": float("inf"),
            }
        )

    return sorted(grouped, key=lambda row: (row["group_order_key"], row["order_key"]))


DECODE_BREAKDOWN_HEADER = [
    "layer_group",
    "layer_count",
    "module/tag",
    "kernel",
    "stream_id",
    "time_us_per_layer",
    "percent_of_current_layer",
    "percent_of_decode_step",
]
XLSX_KERNEL_DISPLAY_LIMIT = 120


def decode_breakdown_values(
    rows: list[dict[str, Any]], full_decode_us: float
) -> list[list[Any]]:
    """Flatten breakdown rows into sheet rows.

    Two shares, both kept as fractions and rendered ``xx.xxx%`` by the writers:

    ``percent_of_current_layer`` divides by the row's LAYER TOTAL — what the
    operator costs within one layer. Empty for rows outside any layer group.

    ``percent_of_decode_step`` scales ``time_us`` (one layer's average) back up
    by ``layer_count`` first, so it is what the whole group costs the step, not
    what one of its layers does.
    """
    values: list[list[Any]] = []
    for row in rows:
        layer_count = int(row.get("layer_count", 1))
        if row.get("suppress_numbers"):
            values.append(
                [row["layer_group"], "", row["module"], row["kernel"], "", "", "", ""]
            )
            continue
        group_us = float(row["time_us"]) * layer_count
        layer_total_us = row.get("layer_total_us")
        in_layer = (
            float(row["time_us"]) / layer_total_us
            if layer_total_us
            else ""  # non_layer / unmatched rows have no owning layer
        )
        values.append(
            [
                row["layer_group"],
                layer_count,
                row["module"],
                row["kernel"],
                row["stream_no"],
                float(row["time_us"]),
                in_layer,
                group_us / full_decode_us if full_decode_us > 0 else 0.0,
            ]
        )
    return values


def write_decode_csv(path: str, values: list[list[Any]]) -> None:
    """Write CSV with repeated group/module cells blanked for readability."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(DECODE_BREAKDOWN_HEADER)
        prev_layer_group: str | None = None
        prev_module: str | None = None
        for row in values:
            layer_group = str(row[0])
            module = str(row[2])
            repeats_group = layer_group == prev_layer_group
            display_layer_group = "" if repeats_group else layer_group
            display_layer_count = "" if repeats_group else row[1]
            display_module = "" if repeats_group and module == prev_module else module
            writer.writerow(
                [
                    display_layer_group,
                    display_layer_count,
                    display_module,
                    row[3],
                    row[4],
                    f"{row[5]:.3f}" if row[5] != "" else "",
                    f"{row[6]:.3%}" if row[6] != "" else "",
                    f"{row[7]:.3%}" if row[7] != "" else "",
                ]
            )
            prev_layer_group = layer_group
            prev_module = module


def merge_same_value_runs(
    ws: Any, column: int, start_row: int, end_row: int, key_column: int | None = None
) -> None:
    """Merge runs of equal cells in *column*.

    ``key_column`` defines the run boundaries when they must not be taken from
    *column* itself — layer_count repeats across unrelated groups, so it merges
    on the layer_group runs instead of its own.
    """
    key_column = key_column or column
    run_start = start_row
    prev_value = ws.cell(row=start_row, column=key_column).value
    for row in range(start_row + 1, end_row + 2):
        value = ws.cell(row=row, column=key_column).value if row <= end_row else None
        if value != prev_value:
            if prev_value not in (None, "") and row - run_start > 1:
                ws.merge_cells(
                    start_row=run_start,
                    start_column=column,
                    end_row=row - 1,
                    end_column=column,
                )
            run_start = row
            prev_value = value


def write_decode_xlsx(path: str, values: list[list[Any]]) -> None:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "decode_breakdown"
    ws.append(DECODE_BREAKDOWN_HEADER)
    for value_row in values:
        display_row = list(value_row)
        kernel = str(display_row[3])
        if len(kernel) > XLSX_KERNEL_DISPLAY_LIMIT:
            display_row[3] = kernel[: XLSX_KERNEL_DISPLAY_LIMIT - 3] + "..."
        ws.append(display_row)
        if len(kernel) > XLSX_KERNEL_DISPLAY_LIMIT:
            ws.cell(row=ws.max_row, column=4).comment = Comment(kernel, "ATOM")

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    end_row = ws.max_row
    if end_row >= 2:
        # Column 3 (module) merges only within a run of the same layer_group,
        # so it is keyed on the pair rather than merged on its own.
        module_run_start = 2
        prev_key = (ws.cell(row=2, column=1).value, ws.cell(row=2, column=3).value)
        for row in range(3, end_row + 2):
            key = (
                ws.cell(row=row, column=1).value if row <= end_row else None,
                ws.cell(row=row, column=3).value if row <= end_row else None,
            )
            if key != prev_key:
                if prev_key[1] not in (None, "") and row - module_run_start > 1:
                    ws.merge_cells(
                        start_row=module_run_start,
                        start_column=3,
                        end_row=row - 1,
                        end_column=3,
                    )
                module_run_start = row
                prev_key = key

        # layer_count first: merging a column blanks every cell but its
        # top-left, so column 1 has to stay intact while it is used as the key.
        merge_same_value_runs(ws, 2, 2, end_row, key_column=1)
        merge_same_value_runs(ws, 1, 2, end_row)

    for row in ws.iter_rows(min_row=2):
        is_total_row = row[3].value == "LAYER TOTAL"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        row[5].number_format = "0.000"
        # Excel percent format: the cell holds the fraction and renders xx.xxx%,
        # so the value stays a real number for sorting and further math.
        row[6].number_format = "0.000%"
        row[7].number_format = "0.000%"

    widths = {
        1: 24,
        2: 12,
        3: 72,
        4: 72,
        5: 10,
        6: 16,
        7: 22,
        8: 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)


def write_decode_output(
    path: str,
    rows: list[dict[str, Any]],
    full_decode_us: float,
    print_head: bool = False,
) -> None:
    stream_map = remap_streams([{"args": {"stream": row["stream"]}} for row in rows])
    breakdown_rows = build_grouped_breakdown_rows(
        rows, stream_map, print_head=print_head
    )
    values = decode_breakdown_values(breakdown_rows, full_decode_us)
    if path.lower().endswith(".xlsx"):
        write_decode_xlsx(path, values)
    else:
        write_decode_csv(path, values)


def main() -> None:
    parser = argparse.ArgumentParser(description="Formal ATOM trace parser")
    parser.add_argument("run_trace")
    parser.add_argument("--capture-trace", default=None)
    parser.add_argument(
        "--output",
        default="decode_breakdown.xlsx",
        help="Output path, .xlsx or .csv (default: decode_breakdown.xlsx).",
    )
    parser.add_argument(
        "--kernel-num",
        type=int,
        default=100,
        help="Number of warmup kernel mappings to print (default: 100).",
    )
    parser.add_argument(
        "--print-head",
        action="store_true",
        help="Include non-layer prologue/epilogue rows in the output.",
    )
    args = parser.parse_args()

    run_events = load_events(args.run_trace)

    # The batch size selects which capture file to open, so it has to be read
    # from the run trace first.
    decode = find_first_decode(run_events)
    batch_size, graph_batch_size = decode_batch_sizes(decode)
    q_len = decode_query_len(decode, batch_size)
    capture_trace = args.capture_trace or resolve_capture_trace(
        args.run_trace, graph_batch_size, q_len
    )
    capture_events = load_events(capture_trace)

    graph = find_capture_graph_for_bs(capture_events, graph_batch_size)
    warmup_start, warmup_end = warmup_window_for_graph(capture_events, graph)
    counts = count_events_in_window(capture_events, warmup_start, warmup_end)

    print(f"Run trace: {args.run_trace}")
    print(f"Capture trace: {capture_trace}")
    print("")
    print("First decode:")
    print(f"  name: {decode.get('name')}")
    print(f"  ts: {decode.get('ts'):.3f}")
    print(f"  dur: {decode.get('dur'):.3f}")
    if graph_batch_size != batch_size:
        print(f"  batch size: {batch_size} (padded to graph bs={graph_batch_size})")
    else:
        print(f"  batch size: {batch_size}")
    print(f"  query len: {q_len if q_len is not None else '?'}")
    print()
    print("Matching capture graph:")
    print(f"  name: {graph.get('name')}")
    print(f"  ts: {graph.get('ts'):.3f}")
    print(f"  dur: {graph.get('dur'):.3f}")
    print("")
    print("Decode warmup window:")
    print(f"  start: {warmup_start:.3f}")
    print(f"  end: {warmup_end:.3f}")
    print(f"  dur: {warmup_end - warmup_start:.3f}")
    print(f"  events: {counts}")
    warmup_mapping = build_warmup_mapping(capture_events, warmup_start, warmup_end)
    print(f"  mapping entries: {len(warmup_mapping)}")
    print_first_warmup_mappings(warmup_mapping, limit=args.kernel_num)

    decode_start, decode_end = decode_gpu_window(run_events, decode)
    replay_kernels = replay_kernels_in_window(run_events, decode_start, decode_end)
    matched_rows = match_replay_to_warmup(replay_kernels, warmup_mapping)
    full_decode_us = decode_end - decode_start
    write_decode_output(
        args.output,
        matched_rows,
        full_decode_us,
        print_head=args.print_head,
    )
    unmatched = sum(1 for row in matched_rows if row["cpu_module"] == "<unmatched>")
    print("")
    print("Decode replay mapping:")
    print(f"  replay kernels: {len(replay_kernels)}")
    print(f"  unmatched kernels: {unmatched}")
    print(f"  output written to: {args.output}")


if __name__ == "__main__":
    main()
